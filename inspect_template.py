import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('du-toan/FILE MẪU.xlsx')

with open('du-toan/inspect_results.txt', 'w', encoding='utf-8') as f:
    f.write(f"Sheets: {wb.sheetnames}\n\n")
    
    for sheet_name in wb.sheetnames:
        f.write("="*80 + "\n")
        f.write(f"SHEET: {sheet_name}\n")
        f.write("="*80 + "\n")
        ws = wb[sheet_name]
        
        # Dimensions
        f.write(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")
        f.write(f"Merged cells: {list(ws.merged_cells.ranges)}\n\n")
        
        # Column widths
        cols_width = {}
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            cols_width[col_letter] = ws.column_dimensions[col_letter].width
        f.write(f"Column widths: {cols_width}\n\n")
        
        # Print grid of values and formatting for first 40 rows
        for r in range(1, min(41, ws.max_row + 1)):
            row_vals = []
            row_styles = []
            for c in range(1, min(16, ws.max_column + 1)):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                val_str = str(val) if val is not None else ""
                row_vals.append(f"{get_column_letter(c)}{r}: {val_str!r}")
                
                # Check style
                style_desc = []
                if cell.font:
                    font = cell.font
                    if font.bold or font.name != 'Arial' or font.size != 11 or font.color:
                        style_desc.append(f"Font({font.name}, {font.size}, B={font.bold}, C={font.color.rgb if font.color else None})")
                if cell.fill and cell.fill.fill_type:
                    fill = cell.fill
                    style_desc.append(f"Fill(type={fill.fill_type}, fg={fill.fgColor.rgb if fill.fgColor else None})")
                if cell.alignment:
                    align = cell.alignment
                    style_desc.append(f"Align(h={align.horizontal}, v={align.vertical})")
                if cell.border:
                    b = cell.border
                    style_desc.append(f"Border(L={b.left.style if b.left else None}, R={b.right.style if b.right else None}, T={b.top.style if b.top else None}, B={b.bottom.style if b.bottom else None})")
                
                if style_desc:
                    row_styles.append(f"{get_column_letter(c)}{r}: {' '.join(style_desc)}")
            
            f.write(f"Row {r} Values: {', '.join(row_vals)}\n")
            if row_styles:
                f.write(f"Row {r} Styles: {'; '.join(row_styles)}\n")
            f.write("\n")
        f.write("\n\n")
